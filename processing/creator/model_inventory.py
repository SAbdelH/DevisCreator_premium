from pathlib import Path
import openpyxl
from processing.enumerations import LevelCritic as LVL

def ModelImportInventory(self):
    try:
        file = Path(self.maindialog.GetOpenDialogFolderPath()) / "model inventaires.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "inventaires"

        for i, name in enumerate(self.LIST.MODEL_INVENTORY_COLUMNS):
            cellule = ws.cell(row=1, column=i + 1)
            cellule.value = name
            cellule.alignment = self.alignerCentrer

        largCol = {1: 43, 3: 27, 6: 15, 7: 16, 8: 16, 9: 16, 10: 16, 11: 16, 12: 71}
        for col in range(1, 13):
            ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = largCol.get(col, 13)

        # validation de données
        __params = {"ws": ws, "col": 6, "debut_ligne": 2, "fin_ligne": 2, "options": ["€", "%"]}
        self.ajouter_validation_liste(**__params)
        __params = {"ws": ws, "col": 7, "debut_ligne": 2, "fin_ligne": 2, "options": ["Oui", "Non"]}
        self.ajouter_validation_liste(**__params)
        __params = {"ws": ws, "col": 8, "debut_ligne": 2, "fin_ligne": 2, "options": ["Oui", "Non"]}
        self.ajouter_validation_liste(**__params)
        __params = {"ws": ws, "col": 9, "debut_ligne": 2, "fin_ligne": 2, "options": ["Achat", "Non"]}
        self.ajouter_validation_liste(**__params)

        # Mise en forme en Tableau
        __params = {"worksheet": ws, "tableName": "Tableau_inventaire", "styleName": "PivotStyleLight3", "add_one": True}
        self.mise_forme_tableau(**__params)

        wb.save(file)
        self.maindialog.show_notification(
            f"Export de {file.as_posix()} fini !", LVL.success
        )
        self.OpenFile(file.as_posix())
    except Exception as e:
        self.maindialog.show_notification(
            str(e), LVL.warning
        )