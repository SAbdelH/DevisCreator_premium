import subprocess
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

from processing.enumerations import TaillePapier as T
from processing.decrypt import APPLESCRIPT



class Excel:
    def convertXlsxToPDF(self, Fichier, pageName=None):
        """
        Converti un Fichier Excel en PDF avec windows et mac
        :param Fichier: Chemin vers le fichier source
        :param pageName: Nom de la feuille pour l'APPLESCRIPT
        :return: Chemin vers le fichier PDF
        """
        pdfPath = None
        if self.systeme == "Windows":
            from win32com.client import Dispatch

            xlsx = Path(Fichier).as_posix()
            pdfPath = Path(xlsx).with_suffix(".pdf")
            excel = Dispatch("Excel.Application")
            excel.Visible = False
            try:
                wb = excel.Workbooks.Open(xlsx)
                wbk = openpyxl.load_workbook(xlsx)
                wb.WorkSheets(wbk.sheetnames).Select()
                wb.ActiveSheet.ExportAsFixedFormat(0, pdfPath)
            except Exception as e:
                self.updateLog(
                    "Le programme n'a pas pu convertir en PDF",
                    titre="non",
                    color="#E74C3C",
                )
                pdfPath = None
            else:
                pass
            finally:
                wb.Close()
                excel.Quit()
                wbk.close()
        elif self.systeme == "Darwin":
            try:
                script_path = Path().home() / "Documents" / "exceltopdf.scpt"
                xlsx = Path(Fichier).as_posix()
                pdfPath = Path(xlsx).with_suffix(".pdf")
                APPLESCRIPT_content = (
                    APPLESCRIPT.replace("<@ExcelPath>", xlsx)
                    .replace("<@PDFPath>", pdfPath.as_posix())
                    .replace("<@SheetName>", pageName)
                )
                with script_path.open("w") as script_file:
                    script_file.write(APPLESCRIPT_content.strip())

                subprocess.run(["osascript", str(script_path)], check=True)
                script_path.unlink(missing_ok=True)
            except subprocess.CalledProcessError as e:
                self.forceSuppression(script_path)
                pdfPath = None
        return pdfPath

    def mise_forme_tableau(self, **kwargs):
        """
        Cette méthode vise à créer une mise en forme tableau
        :key worksheet: La feuille de calcul
        :key tableName : Nom du tableau (Optionnel)
        :key styleName : Nom du style de table
        :key add_one : Booléen pour ajouter plus 1 si maximum col et ligne est égale à 1
        [‘PivotStyleLight5’, ‘TableStyleMedium24’, ‘TableStyleMedium25’, ‘PivotStyleMedium19’, ‘PivotStyleMedium18’,
        ‘PivotStyleMedium13’, ‘PivotStyleMedium12’, ‘PivotStyleMedium11’, ‘PivotStyleMedium10’, ‘PivotStyleMedium17’,
        ‘PivotStyleMedium16’, ‘PivotStyleMedium15’, ‘PivotStyleMedium14’, ‘PivotStyleLight22’, ‘PivotStyleLight23’,
        ‘PivotStyleLight4’, ‘PivotStyleLight21’, ‘PivotStyleLight26’, ‘PivotStyleLight27’, ‘PivotStyleLight24’,
        ‘PivotStyleLight25’, ‘PivotStyleLight28’, ‘PivotStyleLight2’, ‘PivotStyleDark9’, ‘PivotStyleDark6’,
        ‘PivotStyleDark7’, ‘PivotStyleDark4’, ‘PivotStyleDark5’, ‘PivotStyleDark2’, ‘PivotStyleDark3’,
        ‘PivotStyleDark1’, ‘TableStyleMedium22’, ‘TableStyleMedium23’, ‘TableStyleMedium20’, ‘TableStyleMedium21’,
        ‘TableStyleMedium26’, ‘TableStyleMedium27’, ‘TableStyleLight5’, ‘TableStyleLight4’, ‘TableStyleLight9’,
        ‘TableStyleLight8’, ‘PivotStyleLight17’, ‘PivotStyleLight16’, ‘PivotStyleLight15’, ‘PivotStyleLight14’,
        ‘PivotStyleLight13’, ‘PivotStyleLight12’, ‘PivotStyleLight11’, ‘PivotStyleLight10’, ‘PivotStyleLight19’,
        ‘PivotStyleLight18’, ‘PivotStyleLight9’, ‘PivotStyleDark10’, ‘PivotStyleDark11’, ‘PivotStyleDark12’,
        ‘PivotStyleDark13’, ‘PivotStyleDark14’, ‘PivotStyleDark15’, ‘PivotStyleDark16’, ‘PivotStyleDark17’,
        ‘PivotStyleDark18’, ‘PivotStyleDark19’, ‘TableStyleMedium3’, ‘TableStyleMedium2’, ‘TableStyleMedium1’,
        ‘TableStyleMedium7’, ‘TableStyleMedium6’, ‘TableStyleMedium5’, ‘TableStyleMedium4’, ‘TableStyleMedium9’,
        ‘TableStyleMedium8’, ‘PivotStyleLight20’, ‘TableStyleMedium19’, ‘TableStyleMedium18’, ‘TableStyleMedium13’,
        ‘TableStyleMedium12’, ‘TableStyleMedium11’, ‘TableStyleMedium10’, ‘TableStyleMedium17’, ‘TableStyleMedium16’,
        ‘TableStyleMedium15’, ‘TableStyleMedium14’, ‘TableStyleLight20’, ‘TableStyleLight21’, ‘TableStyleDark8’,
        ‘TableStyleDark9’, ‘TableStyleDark6’, ‘TableStyleDark7’, ‘TableStyleDark4’, ‘TableStyleDark5’, ‘TableStyleDark2’,
        ‘TableStyleDark3’, ‘TableStyleDark1’, ‘PivotStyleDark8’, ‘PivotStyleMedium22’, ‘PivotStyleMedium23’,
        ‘PivotStyleMedium20’, ‘PivotStyleMedium21’, ‘PivotStyleMedium26’, ‘PivotStyleMedium27’, ‘PivotStyleMedium24’,
        ‘PivotStyleMedium25’, ‘PivotStyleMedium28’, ‘PivotStyleLight8’, ‘TableStyleLight17’, ‘TableStyleLight16’,
        ‘TableStyleLight15’, ‘TableStyleLight14’, ‘TableStyleLight13’, ‘TableStyleLight12’, ‘TableStyleLight11’,
        ‘TableStyleLight10’, ‘TableStyleLight19’, ‘TableStyleLight18’, ‘PivotStyleDark28’, ‘PivotStyleDark21’,
        ‘PivotStyleDark20’, ‘PivotStyleDark23’, ‘PivotStyleDark22’, ‘PivotStyleDark25’, ‘PivotStyleDark24’,
        ‘PivotStyleDark27’, ‘PivotStyleDark26’, ‘PivotStyleLight1’, ‘TableStyleMedium28’, ‘PivotStyleLight3’,
        ‘TableStyleDark10’, ‘TableStyleDark11’, ‘PivotStyleMedium3’, ‘PivotStyleMedium2’, ‘PivotStyleMedium1’,
        ‘PivotStyleMedium7’, ‘PivotStyleMedium6’, ‘PivotStyleMedium5’, ‘PivotStyleMedium4’, ‘PivotStyleLight7’,
        ‘PivotStyleMedium9’, ‘PivotStyleMedium8’, ‘TableStyleLight3’, ‘PivotStyleLight6’, ‘TableStyleLight2’,
        ‘TableStyleLight1’, ‘TableStyleLight7’, ‘TableStyleLight6’])
        :return:
        """
        worksheet = kwargs.get('worksheet')
        styleName = kwargs.get('styleName', "TableStyleMedium9")
        tableName = kwargs.get('tableName', "Tableau_1")
        add_one =  kwargs.get('add_one', False)

        # 1. Ajouter un tableau si les données sont déjà présentes
        table_range = self.generer_range_feuille(worksheet, add_one)

        # 2. Créer un tableau
        table = Table(displayName=tableName, ref=table_range)

        # 3. Appliquer un style au tableau
        style = TableStyleInfo(
            name=styleName,  # Style prédéfini
            showFirstColumn=False,  # Mettre en évidence la première colonne
            showLastColumn=False,  # Mettre en évidence la dernière colonne
            showRowStripes=True,  # Appliquer des bandes alternées aux lignes
            showColumnStripes=False  # Appliquer des bandes alternées aux colonnes
        )
        table.tableStyleInfo = style

        # 4. Ajouter le tableau à la feuille
        worksheet.add_table(table)

        # 5. Zoom de la feuille
        worksheet.sheet_view.zoomToFit = True
        worksheet.sheet_view.zoomScale = 102

    def mise_en_page(self, **kwargs):
        """
        Création d'une mise en page pour une impression.
        :key wb (Workbook): Instance du classeur à utiliser.
        :key ws (Worksheet): Feuille de calcul active pour la mise en page.
        :key sauvegarde (Path): Le chemin de sauvegarde.
        :key area (str): Plage de cellules à mettre en page (ex: 'A1:D10').
        :key orientation (str): Orientation de la page ('portrait' ou 'paysage').
        :key margin (dict): Marges de la page en pouces.
        :key grillage (bool): Indique si la grille doit être affichée.
        :key apercu (str): Mode d'aperçu avant impression.
        :key zoom (int): Niveau de zoom pour l'affichage.
        :key largPage (float): Nombre de pages en largeur.
        :key hautPage (float): Nombre de pages en hauteur.
        :key PAPERSIZE (TaillePapier): Taille du papier (ex: 'A4', 'Letter').
        :key cheight (float): Hauteur personnalisée de la page.
        :key cwidth (float): Largeur personnalisée de la page.
        :key HeadCenter (str): Texte de l'en-tête centré.
        :key HeadRight (str): Texte de l'en-tête aligné à droite.

        """
        # configuration des parametres
        wb = kwargs["wb"]
        ws = kwargs["ws"]
        sauvegarde = kwargs["sauvegarde"]
        margin = kwargs.get("margin", self.JSON.MARGINF)
        largPage = kwargs.get("largPage", 1)
        hautPage = kwargs.get("hautPage", 1)
        # configuration des tailles papier
        taille = {t: getattr(ws, t.value) if t.name != "CUSTOM" else t.value for t in T}

        ws.page_setup.paperSize = taille.get(kwargs.get("PAPERSIZE", T.A4))
        ws.page_setup.orientation = (
            ws.ORIENTATION_PORTRAIT
            if kwargs.get("orientation", "portrait") == "portrait"
            else ws.ORIENTATION_LANDSCAPE
        )
        ws.print_area = kwargs.get("area", "A1:B1")
        ws.page_margins = openpyxl.worksheet.page.PageMargins(**margin)
        ws.sheet_view.showGridLines = kwargs.get("grillage", True)
        if kwargs.get("PAPERSIZE", T.A4) == T.CUSTOM:
            ws.page_setup.paperHeight = f"{kwargs.get('cheight', 420)}mm"
            ws.page_setup.paperWidth = f"{kwargs.get('cwidth', 297)}mm"
        ws.page_setup.fitToWidth = False if not isinstance(largPage, int) else largPage
        ws.page_setup.fitToHeight = False if not isinstance(hautPage, int) else hautPage
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True
        # Configuration du centrage lors de l'impression
        ws.print_options.horizontalCentered = True  # Centrage horizontal
        # ws.print_options.verticalCentered = True    # Centrage vertical
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.sheet_view.view = kwargs.get(
            "apercu", "normal"
        )  # 'normal, "pageBreakPreview" , 'pageLayout'
        ws.sheet_view.zoomToFit = True
        # Recherche la colonne maximum
        _, _, MC, _ = openpyxl.utils.range_boundaries(kwargs.get("area", "A1:B1"))
        ws.column_dimensions[openpyxl.utils.get_column_letter(MC)].width = 3
        if kwargs.get("zoom"):
            ws.sheet_view.zoomScale = kwargs.get("zoom")
        wb.save(sauvegarde)

    def pied_page(self, **kwargs):
        """
        insertion d'un pied de page Excel
        :param wb: Workbook
        :param ws: Worksheet
        :param sauvegarde: sauvegarde du fichier
        :param gauche: gauche de page
        :param droite: droite de page
        :param centre: centre de page
        :return:
        """
        wb: Workbook = kwargs["wb"]
        ws: Worksheet = kwargs["ws"]
        gauche = kwargs.get("gauche", "&P sur &N")
        centre = kwargs.get("centre", self.TEXT.FOOTER_FACTURE)
        droite = kwargs.get("droite")
        ws.oddFooter.center.text = centre
        ws.oddFooter.left.text = gauche
        ws.oddFooter.right.text = droite
        ws.oddHeader.size = 16
        ws.oddHeader.font = "Arial,Bold"
        wb.save(kwargs.get("sauvegarde"))

    def CentrerMultipleCols(self, workbook: Workbook, worksheet: Worksheet, cell_range: str, sauvegarde: Path):
        """
        Center sur plusieur colonnes du worksheet

        :param workbook: Classeur
        :param worksheet: Feuille
        :param cell_range: Plage des cellules à centrer le texte
        :param sauvegarde: chemin du fichier de sauvegarde
        :return: None
        """
        for row in worksheet[cell_range]:
            for cell in row:
                alignment = cell.alignment.copy()
                alignment.horizontal = "centerContinuous"
                cell.alignment = alignment
        workbook.save(sauvegarde)

    def BordureExterieur(self, workbook, worksheet, bordure, limin: int, limax: int, colmin: int, colmax: int, sauvegarde: str,):
        """
        Donne une bordure exérieur sans effacer le style
        :param workbook: Claseur
        :param worksheet: Feuille
        :param bordure: Type de bordure
        :param limin: numero de ligne minimum
        :param limax: numero de ligne maximum
        :param colmin: numero de colonnes minimum
        :param colmax: numero de colonnes maximum
        :param sauvegarde: Fichier de sauvegarde
        :Example:
        >>> pour centrer de A1 à C1 : self.BordureExterieur(wb, ws, thin, 1, 1, 1, 3, "fichier.xlsx")
        :return: None
        """
        for row in worksheet.iter_rows(
            min_row=limin, max_row=limax, min_col=colmin, max_col=colmax
        ):
            for cell in row:
                lettre, ligne = openpyxl.utils.cell.coordinate_from_string(
                    cell.coordinate
                )
                index_colonne = openpyxl.utils.column_index_from_string(lettre)
                b = {
                    "left": bordure if (index_colonne == colmin) else cell.border.left,
                    "top": (
                        bordure
                        if (ligne == limin or limin == limax)
                        else cell.border.top
                    ),
                    "right": (
                        bordure if (index_colonne == colmax) else cell.border.right
                    ),
                    "bottom": (
                        bordure
                        if (ligne == limax or limin == limax)
                        else cell.border.bottom
                    ),
                }
                cell.border = openpyxl.styles.Border(**b)
        workbook.save(sauvegarde)

    def generer_range_feuille(self, ws: Worksheet, add_one: bool = False):
        """
        Génère automatiquement la plage utilisée dans une feuille Excel.

        Args:
            ws (Worksheet): La feuille de calcul active ou sélectionnée.

        Returns:
            str: La plage sous forme "A1:Z100" en fonction des données de la feuille.
        """
        min_row = ws.min_row
        max_row = (ws.max_row + 1 if ws.max_row == 1 and add_one else ws.max_row)
        min_col = ws.min_column
        max_col = (ws.max_column + 1 if ws.max_column == 1 and add_one else ws.max_column)

        # Convertir les indices de colonnes en lettres (ex : 1 -> A, 3 -> C)
        min_col_letter = get_column_letter(min_col)
        max_col_letter = get_column_letter(max_col)

        # Construire la plage
        table_range = f"{min_col_letter}{min_row}:{max_col_letter}{max_row}"
        return table_range

    def ajouter_validation_liste(self, **kwargs):
        """
        Ajoute une validation de liste à une plage de cellules.

        :key worksheet: La feuille de calcul
        :key colonne: La lettre de la colonne (ex: 'B')
        :key debut_ligne: Première ligne de la validation
        :key fin_ligne: Dernière ligne de la validation
        :key options: Liste des options pour la validation directe
        :key plage_reference: Plage de cellules de référence (ex: '=$A$1:$A$5')
        """
        worksheet = kwargs.get('ws')
        colonne = kwargs.get('col')
        debut_ligne = kwargs.get('debut_ligne')
        fin_ligne = kwargs.get('fin_ligne')
        options = kwargs.get('options')
        plage_reference = kwargs.get('plage_reference')

        # Convertir la colonne en lettre si elle est numérique
        if isinstance(colonne, int):
            colonne = openpyxl.utils.cell.get_column_letter(colonne)

        if options and plage_reference:
            raise ValueError("Spécifiez soit options soit plage_reference, pas les deux")

        if not options and not plage_reference:
            raise ValueError("Spécifiez au moins options ou plage_reference")

        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"' if options else plage_reference,
            allow_blank=True,
            error="Veuillez choisir une option valide",
            errorTitle="Erreur de saisie",
            prompt="Sélectionnez une option dans la liste",
            promptTitle="Liste de choix"
        )

        worksheet.add_data_validation(validation)
        validation.add(f'{colonne}{debut_ligne}:{colonne}{fin_ligne}')