from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from processing.database.model_public import Entreprise

class ExcelPlans:
    def ModelFacture(self, wb: Workbook= None, ws: Worksheet = None, sauvegarde: Path = None):
        wb = openpyxl.Workbook() if not wb else wb
        ws = wb.active if not ws else ws

        COULEUR = {"vert80": "D6F1E8", "bleu-vert-foncéClaire10": "007E8C", "bleu-vert-foncé": "005964"}
        # Largeurs de colonnes
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = (
                self.JSON.LARG_COL_FACTURE.get(col, 26)
            )

        # Hauteurs des lignes
        ws.row_dimensions[1].height = 61
        ws.row_dimensions[13].height = 30

        # Remplissage et couleur Police
        for col in range(1, 8):
            for row in range(1, 14):
                C = ("vert80" if row <= 5 and col <= 4 else "bleu-vert-foncéClaire10"
                if row <= 5 and col > 4 else "bleu-vert-foncé" if row == 13 else "")
                color = COULEUR.get(C, "FFFFFF")
                ws.cell(row, col).fill = openpyxl.styles.PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                # Police
                fontcolor = "FFFFFF" if row == 13 or (col > 4 and row <=5) else "000000"
                fontsize = 24 if (row == 1 and col <= 4) else 18 if ( row == 1 and col == 7 ) else 11
                fontBold = (row==13) or (col>4 and row <=5) or (col==5 and row ==7) or (col==1 and row == 11) or (col==2 and row == 1)
                ws.cell(row, col).font = openpyxl.styles.Font(
                    bold=fontBold, name="Arial", size=fontsize, color=fontcolor
                )

        # Alignement dans les cellules
        ws["B1"].alignment = self.alignerCentrer
        ws["G1"].alignment = self.alignerCentrer

        # fusion des cellules
        column_merge = ["B1:D1", "B2:D2", "B3:D3", "B4:D4", "E2:F2", "E3:F3", "E4:F4", "F7:G7", "F8:G8",
                        "F8:G9","B11:G11", "F7:G9"]
        for plage in column_merge:
            self.CentrerMultipleCols(wb, ws, plage, sauvegarde)


        # INSERT TEXT
        TypeText = self.InvoicePage if self.InvoicePage == "devis" else "Facture"
        ws["G1"] = TypeText.upper()
        ws["E2"] = "N° de facture :"
        ws["E3"] = "Date de facturation :"
        ws["E4"] = "Échéance :"
        ws["E7"] = "Client :"
        ws["E7"].alignment = self.alignerCentrerDroite
        ws["A11"] = "Objet :"
        ws["A11"].alignment = self.alignerCentrerDroite
        # Titres de la facture
        for i, (texte, plage) in enumerate(self.JSON.ENTETE_FACTURE.items()):
            cellule = plage.split(":")[0]
            ws[cellule].value = texte
            if cellule == "A13":
                ws[cellule].alignment = self.alignerCentrerGauche
            else :
                ws[cellule].alignment = self.alignerCentrer
        # Information entreprise
        bold = InlineFont(b=True)
        b = TextBlock
        with self.Session() as session:
            entreprise = session.query(Entreprise).first()
            ws["B1"] = entreprise.nom
            ws["B2"] = entreprise.adresse
            ws["B3"] = f"{str(entreprise.code_postal)}, {entreprise.commune.capitalize()} ({entreprise.departement.capitalize()})"
            ws["B4"] = CellRichText(b(bold, 'mail ✉️ : '), entreprise.mail)
            ws["B5"] = CellRichText(b(bold, 'Téléphone 📞 : '), str(entreprise.telephone))
            ws["C5"] = CellRichText(b(bold, ' /Portable 📱 : '), str(entreprise.portable))
            ws["B5"].alignment = self.alignerCentrerDroite
            ws["C5"].alignment = self.alignerCentrerGauche

        # insertion des images
        img = openpyxl.drawing.image.Image(self.maindialog.images.get("company"))
        img.width = 164.16
        img.height = 116.16
        ws.add_image(img, "A1")
        wb.save(sauvegarde)
