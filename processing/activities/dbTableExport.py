import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from pathlib import Path
from processing.enumerations import LevelCritic as LVL


def dbTableToExcel(self):
    # Obtenir l'élément sélectionné dans le TreeWidget
    current_item = self.maindialog._trw_db_structure.currentItem()
    if current_item and current_item.parent():
        table_name = current_item.text(0)
    else:
        table_name = "table"

    # Définir le nom et chemin par défaut du fichier Excel
    default_name = f"export_{table_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    file_path = Path(self.outputfolder) / "Export" / default_name

    try:
        # Créer le répertoire Export s'il n'existe pas
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name

        # Récupérer les en-têtes des colonnes du TableWidget
        headers = []
        for col in range(self.maindialog._tw_select_table.columnCount()):
            header_item = self.maindialog._tw_select_table.horizontalHeaderItem(col)
            header = header_item.text() if header_item else f"Colonne {col + 1}"
            headers.append(header)

            # Écrire les en-têtes dans la première ligne Excel
            cell = ws.cell(row=1, column=col + 1, value=header)

            # Appliquer un style à l'en-tête
            cell.fill = PatternFill(
                start_color="7dcea0", end_color="7dcea0", fill_type="solid"
            )
            cell.font = cell.font.copy(bold=True)
            self.BordureExterieur(wb, ws, self.thin, cell.row, cell.row, cell.col_idx, cell.col_idx, file_path)

        # Vérifier si le TableWidget contient des données
        row_count = self.maindialog._tw_select_table.rowCount()
        if row_count > 0:
            for row in range(row_count):
                for col in range(self.maindialog._tw_select_table.columnCount()):
                    item = self.maindialog._tw_select_table.item(row, col)
                    value = item.text() if item else ""  # Valeur vide si aucune donnée
                    ws.cell(row=row + 2, column=col + 1, value=value)
                    self.BordureExterieur(wb, ws, self.thin, row + 2, row + 2, col + 1, col + 1, file_path)

        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Lettre de la colonne (A, B, ...)
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Sauvegarder le fichier Excel
        wb.save(file_path)

        # Notification de succès
        self.maindialog.show_notification(
            f"Les données ont été exportées avec succès dans :\n{file_path}",
            LVL.success
        )

    except Exception as e:
        # Notification d'erreur
        self.maindialog.show_notification(
            f"Une erreur est survenue lors de l'export :\n{str(e)}",
            LVL.critical
        )