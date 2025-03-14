from collections import ChainMap
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import distinct, func

from forms.page import populateInvoiceCreatedList, populateInvoiceCreatedCombo
from processing.database.session import WorkSession
from processing.enumerations import LevelCritic as LVL
from processing.database.model_public import Devis, Factures, Entreprise, Activites, Ui_Update, Clients
from processing.enumerations import TaillePapier as T


class ActivityExport:
    def exportInvoice(self):
        id = self.getNextID
        workbook: Workbook = Workbook()
        worksheet: Worksheet = workbook.active
        worksheet.title = id
        sauvegarde = Path(self.outputfolder) / self.InvoicePage.capitalize() / f"{self.InvoicePage.lower()}_{id}.xlsx"
        self.ModelFacture(workbook, worksheet, sauvegarde)
        dlg = self.maindialog
        info_facture = {'client' : dlg._le_invoice_nomclient.text(),
                        'mail_client' : dlg._le_invoice_mailclient.text(),
                        'tel_client' : dlg._le_invoice_numclient.text(),
                        'objet' : dlg._le_invoice_objet.text(),
                        'validDays' : dlg._s_invoice_validity.value(),
                        "expire": (datetime.today() + timedelta(days=dlg._s_invoice_validity.value())).strftime('%d/%m/%Y'),
                        "today": datetime.today().strftime('%d/%m/%Y'),
                        "id": id
                        }
        self.insertInfo(workbook, worksheet, sauvegarde, info_facture)
        self.insertCommand(workbook, worksheet, sauvegarde, info_facture)
        self.finDevis(workbook, worksheet, sauvegarde, info_facture)
        self.decoration(workbook, worksheet, sauvegarde)
        self.convertXlsxToPDF(sauvegarde, worksheet.title)
        populateInvoiceCreatedCombo(self, self.InvoicePage)
        self.maindialog.show_notification(
            f"{sauvegarde.stem} est exporté{'e' if self.InvoicePage == 'Factures' else ''}",
            LVL.success,
        )
        self.setClient(page="invoice")
        populateInvoiceCreatedList(self)

    @property
    def getNextID(self):
        table = Devis if self.InvoicePage.lower() == 'devis' else Factures
        # Dans votre méthode
        with self.Session() as session:
            currentID = session.query(
                distinct(func.substr(table.numero_devis if self.InvoicePage.lower() == 'devis' else table.numero_facture,
                                    1, 10))
            ).filter(
                func.date_part('month', table.crea_date) == func.date_part('month', func.current_date())
            ).count()
        ID = datetime.today().strftime('%m%Y')
        return f"{ID}{currentID + 1:04}"

    def insertInfo(self, wb: Workbook, ws: Worksheet, sauvegarde: str, info: dict):
        ws["G2"] = str(info.get("id", ""))
        ws["G3"] = str(info.get("today", ""))
        ws["G4"] = str(info.get("expire", ""))
        ws["F7"] = str(info.get("client", ""))
        self.CentrerMultipleCols(wb, ws, "F7:G7", sauvegarde)
        ws["F8"] = f"""✉️ {str(info.get("mail_client", ""))}"""
        self.CentrerMultipleCols(wb, ws, "F8:G8", sauvegarde)
        ws["F9"] = f"""📞 {str(info.get("tel_client", ""))}"""
        self.CentrerMultipleCols(wb, ws, "F9:G9", sauvegarde)
        ws["B11"] = str(info.get("objet", ""))
        ws["B11"].alignment = self.alignerCentrerGauche
        ws.merge_cells("B11:G11")

        wb.save(sauvegarde)

    def insertCommand(self, wb: Workbook, ws: Worksheet, sauvegarde: str, info: dict):
        Liste = self.maindialog._lw_list_cart
        total_remise = self.maindialog._ds_invoice_total_remise.value()
        somme = 0
        ORMTable = Devis if self.InvoicePage.lower() == "devis" else Factures
        user = WorkSession.get_current_user()
        columnId = "numero_devis" if self.InvoicePage.lower() == "devis" else "numero_facture"
        contrainte_nom = "devis" if self.InvoicePage.lower() == "devis" else "facture"
        with self.Session() as session:
            for i in range(Liste.count()):
                item = Liste.item(i)
                custom_widget = Liste.itemWidget(item)
                if custom_widget:
                    value = custom_widget.getItemInfo
                    ws.append(
                        [
                            value.get('produit'),
                            "",
                            "",
                            value.get('prix_unite'),
                            value.get('quantite'),
                            (
                                value.get('remise')
                                if value.get('type_remise') == "€"
                                else value.get('remise') / 100
                            ),
                            value.get("prix"),
                        ]
                    )
                    # Ligne impaire
                    row = 14 + i
                    for col in range(1, 8):
                        ws.cell(row=row, column=col).alignment = self.alignerCentrer
                        self.BordureExterieur(wb, ws, self.thin, row, row, col, col, sauvegarde)
                        if row % 2 > 0:
                            ws.cell(row, col).fill = openpyxl.styles.PatternFill(
                                start_color="D6F1E8", end_color="D6F1E8", fill_type="solid"
                            )
                        if col in (4, 6, 7):
                            cell = ws.cell(row=row, column=col)
                            if cell and cell.value > 0.0:
                                typeFormat = (
                                    self.TEXT.FORMAT_EURO
                                    if col in (4, 7) or col == 6 and  value.get('type_remise') == "€"
                                    else self.TEXT.FORMAT_PERCENT
                                )
                                cell.value = float(cell.value)
                                cell.number_format = typeFormat
                                cell.data_type = 'n'
                            else:
                                cell.value = '-'
                    plage = f"A{row}:C{row}"
                    self.CentrerMultipleCols(wb, ws, plage, sauvegarde)

                    somme += float(value.get('prix'))
                    __params = dict(ChainMap(value, info))
                    __params["crea_date"] = func.now()
                    __params["crea_user"] = user.identifiant
                    __params[columnId] = f"{__params.get('id')}_{str(i+1).zfill(2)}"
                    __params["validite"] = int(__params.get("validDays"))
                    __params["total_remise"] = total_remise
                    [__params.pop(cle) for cle in ('dlg', 'icon', 'old_price', 'marque','quantifiable', 'louable',
                                                    'validDays', 'expire', 'today', 'id')
                    ]
                    CT = ORMTable(**__params)
                    session.add(CT)

            client = session.query(Clients).filter(Clients.nom == __params.get('client')).first()
            if client:
                client.commerce = (client.commerce or 0) +(somme - total_remise)

            activite = Activites(
                crea_date=func.now(),
                activites=func.concat(f'{ORMTable.__tablename__}_', func.substr(str(__params[columnId]), 1, 10)),
                action=f'Creation {ORMTable.__tablename__}',
                budget=somme
            )
            updt = Ui_Update(nom=contrainte_nom, crea_date=func.now(), crea_user=func.current_user())
            session.add(activite)
            session.add(updt)
            session.commit()
        wb.save(sauvegarde)

    def finDevis(self, workbook: Workbook, worksheet: Worksheet, sauvegarde: str, info: dict):
        maxRow = worksheet.max_row
        VALIDAY_FACTUREText = self.TEXT.VALIDAY_FACTURE.format(**{"nbjours": str(info.get("validDays"))})
        valable = datetime.today() + timedelta(days=info.get("validDays"))
        thin = self.addStyleAttribute("thin", attributes={"color": "85929e"})
        totalFacture = sum(
            [cell[0].value for cell in worksheet.iter_rows(min_row=14, max_row=maxRow, min_col=7, max_col=7) if
                cell[0].value])
        # insertion des valeurs
        with self.Session() as session:
            company = session.query(Entreprise).first()
            worksheet[f"A{maxRow + 3}"] = self.TEXT.MOD_REGLEMENT_FACTURE
            worksheet[f"A{maxRow + 4}"] = self.TEXT.INFO_BANK_FACTURE.format(**{"bic": company.bic, "iban": company.iban})
            worksheet[f"F{maxRow + 3}"] = "Total HT"
            worksheet[f"G{maxRow + 3}"] = round(float(totalFacture), 2)
            worksheet[f"G{maxRow + 3}"].number_format = self.TEXT.FORMAT_EURO
            worksheet[f"G{maxRow + 3}"].data_type = 'n'

            worksheet[f"F{maxRow + 4}"] = "Total TTC"
            worksheet[f"G{maxRow + 4}"] = round(float(totalFacture), 2)
            worksheet[f"G{maxRow + 4}"].number_format = self.TEXT.FORMAT_EURO
            worksheet[f"G{maxRow + 4}"].data_type = 'n'
            worksheet[f"A{maxRow + 5}"] = VALIDAY_FACTUREText
            worksheet[f"F{maxRow + 5}"] = (
                f"Offre valable jusqu'au {valable.strftime('%d/%m/%Y')}"
            )
            worksheet[f"E{maxRow + 7}"] = "Bon pour accord et signature"
            worksheet[f"E{maxRow + 8}"] = "Fait à :"
            worksheet[f"F{maxRow + 8}"] = company.ville if company.ville else company.commune
            worksheet[f"g{maxRow + 8}"] = f"Le : {datetime.today().strftime('%d/%m/%Y')}"

        # police
        worksheet[f"A{maxRow + 3}"].font = openpyxl.styles.Font(
            bold=True, name="Calibri", size=11
        )
        worksheet[f"F{maxRow + 3}"].font = openpyxl.styles.Font(
            bold=True, name="Calibri", size=12
        )
        worksheet[f"G{maxRow + 3}"].font = openpyxl.styles.Font(
            bold=True, name="Calibri", size=12
        )
        worksheet[f"F{maxRow + 4}"].font = openpyxl.styles.Font(
            bold=True, name="Calibri", size=12, color="ffffff"
        )
        worksheet[f"G{maxRow + 4}"].font = openpyxl.styles.Font(
            bold=True, name="Calibri", size=12, color="ffffff"
        )
        worksheet[f"E{maxRow + 7}"].font = openpyxl.styles.Font(
            bold=False, name="Calibri", size=11, color="5dade2"
        )

        # dimension
        worksheet.row_dimensions[maxRow + 4].height = 43
        worksheet.row_dimensions[maxRow + 5].height = 30
        worksheet.row_dimensions[maxRow + 9].height = 42

        # bordures
        self.BordureExterieur(
            workbook, worksheet, thin, maxRow + 3, maxRow + 5, 1, 3, sauvegarde
        )
        [
            self.BordureExterieur(
                workbook,
                worksheet,
                thin,
                cell.row,
                cell.row,
                cell.col_idx,
                cell.col_idx,
                sauvegarde,
            )
            for ligne in worksheet.iter_rows(
            min_row=maxRow + 3, max_row=maxRow + 5, min_col=6, max_col=7
        )
            for cell in ligne
        ]
        [
            setattr(
                worksheet.cell(maxRow + 8, col),
                "border",
                openpyxl.styles.Border(**{"right": thin}),
            )
            for col in range(5, 8)
        ]
        self.BordureExterieur(
            workbook, worksheet, thin, maxRow + 8, maxRow + 9, 5, 7, sauvegarde
        )

        # remplissage
        [
            setattr(
                worksheet.cell(maxRow + 4, col),
                "fill",
                openpyxl.styles.PatternFill(
                    start_color="005964", end_color="005964", fill_type="solid"
                ),
            )
            for col in range(6, 8)
        ]

        # Alignement
        [
            setattr(
                worksheet.cell(ligne, 1),
                "alignment",
                openpyxl.styles.Alignment(wrap_text=True),
            )
            for ligne in range(maxRow + 3, maxRow + 6)
        ]
        worksheet[f"F{maxRow + 3}"].alignment = self.alignerCentrer
        worksheet[f"F{maxRow + 4}"].alignment = self.alignerCentrer
        worksheet[f"G{maxRow + 3}"].alignment = self.alignerCentrer
        worksheet[f"G{maxRow + 4}"].alignment = self.alignerCentrer
        self.CentrerMultipleCols(
            workbook, worksheet, f"F{maxRow + 5}:G{maxRow + 5}", sauvegarde
        )
        self.CentrerMultipleCols(
            workbook, worksheet, f"E{maxRow + 7}:G{maxRow + 7}", sauvegarde
        )
        worksheet[f"E{maxRow + 8}"].alignment = self.alignerCentrerDroite
        worksheet[f"G{maxRow + 8}"].alignment = self.alignerCentrerDroite

        # fusion des cellules
        worksheet.merge_cells(f"A{maxRow + 3}:C{maxRow + 3}")
        worksheet.merge_cells(f"A{maxRow + 4}:C{maxRow + 4}")
        worksheet.merge_cells(f"A{maxRow + 5}:C{maxRow + 5}")

        workbook.save(filename=sauvegarde)

    def decoration(self, wb: Workbook, ws: Worksheet, sauvegarde: str ):
        with self.Session() as session:
            entreprise = session.query(Entreprise).first()

            _i = {"company": entreprise.nom, "capital": entreprise.capital, "siret": entreprise.siren,
                  "siren": entreprise.siren, "ape": entreprise.code_ape}
            parameters = {
                "wb": wb,
                "ws": ws,
                "sauvegarde": sauvegarde,
                "centre": self.TEXT.FOOTER_FACTURE.format(**_i)
            }
            self.pied_page(**parameters)
            (min_col, min_row, max_col, max_row) = openpyxl.utils.cell.range_boundaries(
                ws.dimensions
            )
            area = f"{openpyxl.utils.cell.get_column_letter(min_col)}{min_row}:{openpyxl.utils.cell.get_column_letter(max_col + 1)}{max_row + 1}"
            parameters = {
                "wb": wb,
                "ws": ws,
                "sauvegarde": sauvegarde,
                "PAPERSIZE": T.A4,
                "area": area,
                "grillage": False,
                "apercu": "pageBreakPreview",
                "hautPage": "auto",
                "margin": self.JSON.MARGINF,
            }
            self.mise_en_page(**parameters)