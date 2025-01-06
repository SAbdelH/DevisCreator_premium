from datetime import date

import openpyxl
from shutil import copy2
from pathlib import Path

from PySide6.QtWidgets import QDialog
from sqlalchemy import func, and_

from forms.gui import AchatDialog
from processing.database.model_private import Chemin
from processing.database.model_public import Inventaires, Activites, Achat, Ui_Update
from processing.database.session import WorkSession
from processing.enumerations import LevelCritic as LVL


class ActivityInsert:
    def inventoryInsert(self, action: str|None = None):
        exist_inv = 0
        MODEL_IMPORT = self.maindialog._le_inventory_import_path.text().strip()
        if Path(MODEL_IMPORT).suffix == ".xlsx":
            wb = openpyxl.load_workbook(MODEL_IMPORT)
            ws = wb.active
            colExist = {
                ws.cell(row=1, column=i).value.strip(): i
                for i in range(1, ws.max_column + 1)
                if ws.cell(row=1, column=i).value
            }

            exist = True
            for i in self.LIST.MODEL_INVENTORY_COLUMNS:
                if i not in colExist:
                    exist = False
            if exist:
                with self.Session() as session:
                    for idx, ligne in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
                        __params = {nameColumn: ligne[index - 1] for nameColumn, index in colExist.items()}
                        __params['quantifiable'] = True if __params.get('quantifiable').lower() == 'oui' else False
                        __params['louable'] = True if __params.get('louable').lower() == 'oui' else False
                        __params['crea_user'] = WorkSession.get_current_user().identifiant
                        exist_inv = exist_inv if (result := self.__action(session, __params)) == 0 else result
                    updt = Ui_Update(nom='inventory', crea_date=func.now(), crea_user=func.current_user())
                    session.add(updt)
                    session.commit()
        else:
            remise = self.maindialog._ds_inventory_remise.value()
            type_remise = self.maindialog._cbx_inventory_type_remise.currentText()
            purchase = self.__achat if action == 'purchase' else 0
            __params = {
                "nom": self.maindialog._le_inventory_name.text().strip(),
                "marque": self.maindialog._le_inventory_marque.text().strip(),
                "prix": self.maindialog._ds_inventory_price.value(),
                "quantite": self.maindialog._s_inventory_quantity.value(),
                "remise": None if remise is None or remise == 0 else remise,
                "type_remise": None if remise == None or remise == 0 else "%" if type_remise == 'En pourcentage' else "€",
                "quantifiable": self.maindialog._cb_inventory_quantifiable.isChecked(),
                "louable": self.maindialog._cb_inventory_location.isChecked(),
                "achat": "Achat" if purchase > 0 else "Non",
                "prix achat": purchase,
                "date_fabric": self.maindialog._de_inventory_fabric.date().toString("yyyy-MM-dd"),
                "lien": self.maindialog._le_inventory_illustration_path.text().strip(),
                "crea_user" : WorkSession.get_current_user().identifiant
            }
            with self.Session() as session:
                exist_inv = exist_inv if (result := self.__action(session, __params, 'dlg')) == 0 else result
                updt = Ui_Update(nom='inventory', crea_date=func.now(), crea_user=func.current_user())
                session.add(updt)
                session.commit()

        self.maindialog.show_notification(
                    f"Les inventaires ont été {'mis-à-jour' if exist_inv != 0 else 'importé'}",
                    LVL.success,
                )

    def __action(self, session, __params: dict, origin: str = 'xlsx'):
        exist_inv = session.query(Inventaires).filter(Inventaires.nom == __params.get('nom')).count()
        if exist_inv == 0:
            self.__insert(session, __params)
        else:
            self.__update(session, __params, origin)
        return exist_inv

    def __insert(self, session, __params: dict):
        if (lien:=__params.get('lien')):
            source = Path(lien.strip())
            self.__copyImage(source, __params.get('nom'))
        __params.pop("lien")
        budget = None
        if (achat:=__params.get('achat'))=="Achat":
            Dachat = Achat(nom=__params['nom'], prix=__params['prix achat'], quantite=__params['quantite'],
                        crea_date=func.now(), crea_user=func.current_user())
            budget = __params['prix achat']
            session.add(Dachat)
        __params.pop("achat")
        __params.pop("prix achat")

        inventaire = Inventaires(**__params)
        activite = Activites(crea_date=func.now(),
                            activites=__params['nom'],
                            action='Achat inventaire' if achat=="Achat" else 'Ajout inventaire',
                            budget=budget)
        session.add(inventaire)
        session.add(activite)
        session.commit()

    def __update(self, session, __params: dict, origin: str = 'xlsx'):
        inventaire = session.query(Inventaires).filter(Inventaires.nom == __params.get('nom')).first()
        if (lien:=__params.get('lien')):
            source = Path(lien.strip())
            self.__copyImage(source, __params.get('nom'))
        __params.pop("lien")

        if origin == 'xlsx' :
            __params["quantite"] = __params['quantite']+inventaire.quantite
        data = {getattr(inventaire, key): value for key, value in __params.items()}
        inventaire.update(data)
        if (achat := __params.get('achat')) == "Achat":
            Dachat = Achat(nom=__params['nom'], prix=__params['prix achat'], quantite=__params['quantite'],
                        crea_date=func.now(), crea_user=func.current_user())
            session.add(Dachat)
        __params.pop("achat")
        __params.pop("prix achat")

        activites = Activites(
            crea_date=date.today(),  # Date du jour
            activites=__params.get('nom'),  # Type d'activité
            action='Mise-à-jour inventaire'  # Action réalisée
        )
        session.add(activites)
        session.commit()

    @property
    def __achat(self):
        dialog = AchatDialog()
        if dialog.exec_() == QDialog.Accepted:
            return dialog.get_price()
        else:
            return 0

    def __copyImage(self, source: Path, nom):
        with self.privateSession() as session:
            inventory_path = (session.query(Chemin.path)
                                .filter(
                                    and_(Chemin.name=='inventaire', Chemin.company==self.GROUP)
                                ).first())
            if inventory_path:
                copy2(src = source, dst = Path(inventory_path[0]) / f"{nom}{source.suffix}")