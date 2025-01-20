import openpyxl

from processing.creator.excel import Excel

e = Excel()
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Facture Test"
bgImg = '/Users/abdelhafidhousoufou/Downloads/Neutral Ripped Paper Business Invoice.png'

img = openpyxl.drawing.image.Image(bgImg)
img.width = 164.16
img.height = 116.16
ws.add_image(img, "A1")

LARG_COL_FACTURE = {4: 16, 5: 13, 6: 13, 7: 19}
COULEUR = {"vert80": "D6F1E8", "bleu-vert-foncéClaire10": "007E8C", "bleu-vert-foncé": "005964"}

# Largeurs de colonnes
for col in range(1, 8):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = (
        LARG_COL_FACTURE.get(col, 26)
    )

# Hauteurs des lignes
ws.row_dimensions[1].height = 61
ws.row_dimensions[13].height = 30

# Remplissage et couleur Police
for col in range(1, 8):
    for row in range(1, 14):
        C = ("vert80" if row <=5 and col <= 4 else "bleu-vert-foncéClaire10"
                if row <=5 and col > 4 else "bleu-vert-foncé" if row == 13 else "")
        color = COULEUR.get(C, "FFFFFF")
        ws.cell(row, col).fill = openpyxl.styles.PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )

# fusion des cellules
column_merge = ["B1:D1", "B2:D2", "B3:D3", "B4:D4", "B5:D5", "E2:F2", "E3:F3", "E4:F4", "F7:G7", "F8:G8", "F8:G9",
                "B11:G11", "F7:G9"]
for plage in column_merge:
    e.CentrerMultipleCols(wb, ws, plage, "test_facture.xlsx")

wb.save("test_facture.xlsx")